# generate_inventory_report.py - FIXED: Added "Intra Gudang Masuk" to AI00
# Group MB51 by (Material, Plant, Storage, Movement Type, Movement Type Text)
# Then map mv_text -> mv_grouping -> determine target column
# FILTER: Only process GS00, BS00, AI00, TR00, EMPTY_STORAGE

import sys
import json
import os
import datetime
import traceback
from collections import defaultdict
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from concurrent.futures import ThreadPoolExecutor
import warnings

def log(msg):
    """Log to stderr"""
    print(f"[worker] {msg}", file=sys.stderr, flush=True)

def find_col(df_cols, candidates):
    """Find first matching column name"""
    lower_map = {str(c).strip().lower(): c for c in df_cols if str(c).strip() and str(c).strip() != 'nan'}
    
    for cand in candidates:
        cand_lower = str(cand).lower()
        if cand_lower in lower_map:
            return lower_map[cand_lower]
    
    for cand in candidates:
        cand_lower = str(cand).lower()
        for key, original in lower_map.items():
            if cand_lower in key:
                return original
    return None

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def get_column_index(letter):
    """Convert Excel column letter to index (1-based)"""
    if len(letter) == 1:
        return ord(letter) - ord('A') + 1
    return sum([(ord(ch) - 64) * (26 ** (len(letter)-i-1)) for i, ch in enumerate(letter)])

def read_sheet(file_path, sheet_name):
    """Read single sheet - parallelizable"""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
            df.columns = [str(c).strip() if not pd.isna(c) else f"Unnamed_{i}" 
                         for i, c in enumerate(df.columns)]
            return sheet_name, df
    except Exception as e:
        log(f"Error reading {sheet_name}: {str(e)}")
        return sheet_name, None

class SheetCache:
    """Cache for sheet lookups - build once, query many times"""
    
    def __init__(self, sheets_dict):
        self.sheets = sheets_dict
        self.caches = {}
        self._build_caches()
    
    def _build_caches(self):
        """Pre-build all lookup caches"""
        log("Building sheet lookup caches...")
        
        # Cache SALDO AWAL
        if 'SALDO AWAL' in self.sheets:
            df = self.sheets['SALDO AWAL']
            cols = list(df.columns)
            mat_col = find_col(cols, ["Kode Material", "Material"])
            plant_col = find_col(cols, ["Plant", "Plnt"])
            sloc_col = find_col(cols, ["Storage Loc", "Storage Location"])
            amt_col = find_col(cols, ["Closing Stock (pcs)", "QTY", "Closing Stock"])
            
            if mat_col and amt_col:
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                df_clean['sloc'] = df_clean[sloc_col].astype(str).str.strip() if sloc_col else ''
                df_clean['amount'] = pd.to_numeric(df_clean[amt_col], errors='coerce').fillna(0)
                
                grouped = df_clean.groupby(['material', 'plant', 'sloc'], dropna=False)['amount'].sum()
                self.caches['saldo_awal'] = grouped.to_dict()
                log(f"  SALDO AWAL cache: {len(self.caches['saldo_awal'])} entries")
        
        # Cache SALDO AWAL MB5B
        if 'SALDO AWAL MB5B' in self.sheets:
            df = self.sheets['SALDO AWAL MB5B']
            
            header_row_idx = None
            if len(df) > 1:
                for i in range(min(10, len(df))):
                    row_values = df.iloc[i].astype(str).str.lower().tolist()
                    if 'material' in ' '.join(row_values):
                        header_row_idx = i
                        break
            
            if header_row_idx is not None:
                df.columns = df.iloc[header_row_idx]
                df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
            
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plnt", "Plant"])
            gs_col = find_col(cols, ["GS"])
            bs_col = find_col(cols, ["BS"])
            
            if mat_col and (gs_col or bs_col):
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                
                if gs_col:
                    df_clean['gs_amount'] = pd.to_numeric(df_clean[gs_col], errors='coerce').fillna(0)
                    grouped_gs = df_clean.groupby(['material', 'plant'], dropna=False)['gs_amount'].sum()
                    self.caches['mb5b_awal_gs'] = grouped_gs.to_dict()
                    log(f"  MB5B AWAL GS cache: {len(self.caches['mb5b_awal_gs'])} entries")
                
                if bs_col:
                    df_clean['bs_amount'] = pd.to_numeric(df_clean[bs_col], errors='coerce').fillna(0)
                    grouped_bs = df_clean.groupby(['material', 'plant'], dropna=False)['bs_amount'].sum()
                    self.caches['mb5b_awal_bs'] = grouped_bs.to_dict()
                    log(f"  MB5B AWAL BS cache: {len(self.caches['mb5b_awal_bs'])} entries")
        
        # Cache MB5B
        if '13. MB5B' in self.sheets:
            df = self.sheets['13. MB5B']
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plnt", "Plant"])
            gs_col = find_col(cols, ["GS"])
            bs_col = find_col(cols, ["BS"])
            
            if mat_col and (gs_col or bs_col):
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                
                if gs_col:
                    df_clean['gs_amount'] = pd.to_numeric(df_clean[gs_col], errors='coerce').fillna(0)
                    grouped_gs = df_clean.groupby(['material', 'plant'], dropna=False)['gs_amount'].sum()
                    self.caches['mb5b_gs'] = grouped_gs.to_dict()
                    log(f"  MB5B GS cache: {len(self.caches['mb5b_gs'])} entries")
                
                if bs_col:
                    df_clean['bs_amount'] = pd.to_numeric(df_clean[bs_col], errors='coerce').fillna(0)
                    grouped_bs = df_clean.groupby(['material', 'plant'], dropna=False)['bs_amount'].sum()
                    self.caches['mb5b_bs'] = grouped_bs.to_dict()
                    log(f"  MB5B BS cache: {len(self.caches['mb5b_bs'])} entries")
        
        # Cache EDS
        if '14. SALDO AKHIR EDS' in self.sheets:
            df = self.sheets['14. SALDO AKHIR EDS']
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plant"])
            sloc_col = find_col(cols, ["Storage Location", "Storage Loc"])
            amt_col = find_col(cols, ["Quantity", "QTY"])
            
            if mat_col and amt_col:
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                df_clean['sloc'] = df_clean[sloc_col].astype(str).str.strip() if sloc_col else ''
                df_clean['amount'] = pd.to_numeric(df_clean[amt_col], errors='coerce').fillna(0)
                
                grouped = df_clean.groupby(['material', 'plant', 'sloc'], dropna=False)['amount'].sum()
                self.caches['eds'] = grouped.to_dict()
                log(f"  EDS cache: {len(self.caches['eds'])} entries")

    def get_saldo_awal(self, material, plant, sloc_type):
        if 'saldo_awal' not in self.caches:
            return 0.0
        return self.caches['saldo_awal'].get((material, plant, sloc_type), 0.0)
    
    def get_mb5b_awal(self, material, plant, sloc_type):
        cache_key = 'mb5b_awal_gs' if sloc_type == "GS" else 'mb5b_awal_bs'
        if cache_key not in self.caches:
            return 0.0
        return self.caches[cache_key].get((material, plant), 0.0)
    
    def get_mb5b(self, material, plant, sloc_type):
        cache_key = 'mb5b_gs' if sloc_type == "GS" else 'mb5b_bs'
        if cache_key not in self.caches:
            return 0.0
        return self.caches[cache_key].get((material, plant), 0.0)
    
    def get_eds(self, material, plant, sloc_type):
        if 'eds' not in self.caches:
            return 0.0
        return self.caches['eds'].get((material, plant, sloc_type), 0.0)

def main():
    try:
        payload = json.load(sys.stdin)
        files = payload.get("files", {})
        master_inventory = payload.get("master_inventory", [])
        master_movement = payload.get("master_movement", [])

        mb51_path = files.get("mb51")
        main_path = files.get("main")

        if not mb51_path or not main_path:
            raise ValueError("Payload must include files.mb51 and files.main paths")

        # Load master data
        log("Loading master data...")
        df_master_inv = pd.DataFrame(master_inventory)
        df_master_mov = pd.DataFrame(master_movement)

        df_master_inv.columns = [str(c).strip().lower() if not pd.isna(c) else f"col_{i}" 
                                  for i, c in enumerate(df_master_inv.columns)]
        df_master_mov.columns = [str(c).strip().lower() if not pd.isna(c) else f"col_{i}" 
                                  for i, c in enumerate(df_master_mov.columns)]

        # Inventory mapping
        log("Creating inventory lookup...")
        df_master_inv['plant'] = df_master_inv['plant'].astype(str).str.strip()
        inv_map = df_master_inv.set_index('plant')[['area', 'kode_dist', 'profit_center']].to_dict('index')
        
        # Build mv_text -> mv_grouping mapping
        log("Building movement text -> grouping mapping...")
        df_master_mov['mv_text'] = df_master_mov['mv_text'].astype(str).str.strip().str.lower()
        df_master_mov['mv_grouping'] = df_master_mov['mv_grouping'].astype(str).str.strip()
        
        # Create mapping: mv_text -> mv_grouping (use first occurrence if duplicates)
        mv_text_to_grouping = {}
        for _, row in df_master_mov.iterrows():
            mv_text = row['mv_text']
            mv_grouping = row['mv_grouping']
            if mv_text and mv_text != 'nan' and mv_grouping and mv_grouping != 'nan':
                if mv_text not in mv_text_to_grouping:
                    mv_text_to_grouping[mv_text] = mv_grouping
        
        log(f"  Created {len(mv_text_to_grouping)} mv_text -> mv_grouping mappings")
        
        # Build storage + mv_grouping -> column mapping
        log("Building (storage, mv_grouping) -> column mapping...")
        storage_grouping_to_column = {
            # GS00
            ("GS00", "Terima Barang"): "R",
            ("GS00", "Retur Beli"): "S",
            ("GS00", "Penjualan"): "T",
            ("GS00", "Retur Jual"): "U",
            ("GS00", "Intra Gudang Masuk"): "V",
            ("GS00", "Intra Gudang"): "W",
            ("GS00", "Transfer Stock"): "X",
            ("GS00", "Pemusnahan"): "Y",
            ("GS00", "Adjustment"): "Z",
            
            # BS00
            ("BS00", "Terima Barang"): "AB",
            ("BS00", "Retur Beli"): "AC",
            ("BS00", "Penjualan"): "AD",
            ("BS00", "Retur Jual"): "AE",
            ("BS00", "Transfer Stock"): "AF",
            ("BS00", "Pemusnahan"): "AG",
            ("BS00", "Adjustment"): "AH",
            ("BS00", "Intra Gudang Masuk"): "AI",
            
            # AI00 - FIXED: Added "Intra Gudang Masuk" as new column AS
            ("AI00", "Terima Barang"): "AJ",
            ("AI00", "Retur Beli"): "AK",
            ("AI00", "Penjualan"): "AL",
            ("AI00", "Retur Jual"): "AM",
            ("AI00", "Intra Gudang Masuk"): "AN",  # NEW COLUMN
            ("AI00", "Intra Gudang"): "AO",
            ("AI00", "Transfer Stock"): "AP",
            ("AI00", "Pemusnahan"): "AQ",
            ("AI00", "Adjustment"): "AR",
            
            # TR00 - Shifted by 1 column
            ("TR00", "Terima Barang"): "AT",
            ("TR00", "Retur Beli"): "AU",
            ("TR00", "Penjualan"): "AV",
            ("TR00", "Retur Jual"): "AW",
            ("TR00", "Intra Gudang"): "AX",
            ("TR00", "Transfer Stock"): "AY",
            ("TR00", "Pemusnahan"): "AZ",
            ("TR00", "Adjustment"): "BA",
            
            # 641/642 (empty storage) - Shifted by 1 column
            ("EMPTY_STORAGE", "Intra Gudang"): "BC",  # 641
            ("EMPTY_STORAGE", "Intra Gudang"): "BD",  # 642
        }
        
        log(f"  Created {len(storage_grouping_to_column)} (storage, mv_grouping) -> column mappings")

        # Read MB51
        log(f"Reading MB51...")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_mb51 = pd.read_excel(mb51_path, sheet_name=0, engine="openpyxl", dtype=str)
        
        df_mb51.columns = [str(c).strip() if not pd.isna(c) else f"Unnamed_{i}" 
                          for i, c in enumerate(df_mb51.columns)]

        # Find columns
        mb_cols = list(df_mb51.columns)
        col_posting = find_col(mb_cols, ["Posting Date"])
        col_material = find_col(mb_cols, ["Material"])
        col_plant = find_col(mb_cols, ["Plant", "Plnt"])
        col_movement = find_col(mb_cols, ["Movement type", "Movement Type"])
        col_movement_text = find_col(mb_cols, ["Movement Type Text"])
        col_amount = find_col(mb_cols, ["Quantity"])
        col_sloc = find_col(mb_cols, ["Storage", "Storage Location", "Storage Loc"])
        col_material_desc = find_col(mb_cols, ["Material description"])

        missing = []
        for name, col in [("Posting Date", col_posting), ("Material", col_material), 
                         ("Plant", col_plant), ("Movement type", col_movement),
                         ("Movement Type Text", col_movement_text), ("Quantity", col_amount)]:
            if not col:
                missing.append(name)
        
        if missing:
            raise ValueError(f"Missing MB51 columns: {', '.join(missing)}")

        # Rename
        rename_dict = {
            col_posting: "posting_date", col_material: "material",
            col_plant: "plant", col_movement: "mv_type",
            col_movement_text: "mv_text", col_amount: "amount"
        }
        if col_sloc:
            rename_dict[col_sloc] = "sloc"
        if col_material_desc:
            rename_dict[col_material_desc] = "material_desc_mb51"
        
        df_mb51 = df_mb51.rename(columns=rename_dict)

        # Convert data types
        log("Converting data types...")
        
        try:
            date_numeric = pd.to_numeric(df_mb51["posting_date"], errors='coerce')
            df_mb51["posting_date"] = pd.to_datetime(
                date_numeric,
                origin='1899-12-30',
                unit='D',
                errors='coerce'
            )
            
            valid_dates = df_mb51["posting_date"].notna().sum()
            log(f"  ✓ Converted {valid_dates}/{len(df_mb51)} dates successfully")
                
        except Exception as e:
            log(f"  ERROR converting dates: {str(e)}")
            df_mb51["posting_date"] = pd.NaT
        
        # Convert amount - keep negative values
        df_mb51["amount"] = pd.to_numeric(df_mb51["amount"], errors="coerce")
        
        neg_count = (df_mb51["amount"] < 0).sum()
        pos_count = (df_mb51["amount"] > 0).sum()
        zero_count = (df_mb51["amount"] == 0).sum()
        nan_count = df_mb51["amount"].isna().sum()
        total_amount = df_mb51["amount"].sum()
        
        log(f"  === AMOUNT DISTRIBUTION (RAW MB51) ===")
        log(f"  Positive: {pos_count}, Negative: {neg_count}, Zero: {zero_count}, NaN: {nan_count}")
        log(f"  Total sum: {total_amount:,.2f}")
        
        df_mb51["plant"] = df_mb51["plant"].astype(str).str.strip()
        df_mb51["mv_type"] = df_mb51["mv_type"].astype(str).str.strip()
        df_mb51["material"] = df_mb51["material"].astype(str).str.strip()
        df_mb51["mv_text"] = df_mb51["mv_text"].astype(str).str.strip().str.lower()
        
        # Handle storage with proper null detection
        if 'sloc' in df_mb51.columns:
            df_mb51["is_empty_storage"] = (
                (df_mb51["sloc"].isna()) |
                (df_mb51["sloc"].isnull()) |
                (df_mb51["sloc"].astype(str).str.strip() == '') |
                (df_mb51["sloc"].astype(str).str.strip().str.upper() == 'NAN') |
                (df_mb51["sloc"].astype(str).str.strip().str.upper() == 'NONE')
            )
            
            df_mb51["storage"] = df_mb51["sloc"].astype(str).str.strip().str.upper()
            df_mb51.loc[df_mb51["is_empty_storage"], "storage"] = 'EMPTY_STORAGE'
        else:
            df_mb51["storage"] = "EMPTY_STORAGE"
            df_mb51["is_empty_storage"] = True
        
        empty_count = df_mb51["is_empty_storage"].sum()
        log(f"  Rows with empty/null storage: {empty_count}/{len(df_mb51)} ({empty_count/len(df_mb51)*100:.1f}%)")

        # Map inventory
        log("Mapping inventory data...")
        df_inv_lookup = pd.DataFrame([
            {'plant': k, **v} for k, v in inv_map.items()
        ])
        df_mb51 = df_mb51.merge(
            df_inv_lookup[['plant', 'area', 'kode_dist', 'profit_center']], 
            on='plant', how='left'
        )

        # Determine report period
        log("Determining report period...")
        if df_mb51["posting_date"].dropna().empty:
            report_month_dt = datetime.datetime.now()
        else:
            report_month_dt = df_mb51["posting_date"].dropna().max()
        
        bulan = report_month_dt.strftime("%B").upper()
        tahun = report_month_dt.year
        prev_month_dt = report_month_dt
        prev_month = prev_month_dt.strftime("%B").upper()
        prev_year = prev_month_dt.year
        bulan_only = bulan

        log(f"  Report period: {bulan} {tahun}")

        # Read main file sheets
        log("Reading main file sheets...")
        required_sheets = ['SALDO AWAL', 'SALDO AWAL MB5B', '13. MB5B', 
                          '14. SALDO AKHIR EDS', 'Output Report INV ARUS BARANG']
        
        sheets_dict = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = [executor.submit(read_sheet, main_path, sheet) for sheet in required_sheets]
            for future in futures:
                sheet_name, df = future.result()
                if df is not None:
                    sheets_dict[sheet_name] = df

        sheet_cache = SheetCache(sheets_dict)

        # Get existing materials
        existing_materials = []
        if 'Output Report INV ARUS BARANG' in sheets_dict:
            df_existing = sheets_dict['Output Report INV ARUS BARANG']
            if df_existing.shape[0] > 8 and df_existing.shape[1] >= 7:
                df_existing_materials = df_existing.iloc[7:, [1, 5]].copy()
                df_existing_materials.columns = ['plant', 'material']
                df_existing_materials = df_existing_materials[
                    (df_existing_materials['material'].notna()) & 
                    (df_existing_materials['material'].astype(str).str.strip() != '') &
                    (df_existing_materials['material'].astype(str).str.strip() != 'nan')
                ]
                
                log("Loading existing materials from main file...")
                for idx, row in df_existing_materials.iterrows():
                    plant = str(row['plant']).strip().upper()
                    material = str(row['material']).strip()
                    
                    if plant in inv_map:
                        existing_materials.append({
                            'material': material, 'plant': plant,
                            'area': inv_map[plant]['area'],
                            'kode_dist': inv_map[plant]['kode_dist'],
                            'profit_center': inv_map[plant]['profit_center']
                        })
                    else:
                        existing_materials.append({
                            'material': material, 'plant': plant,
                            'area': '', 'kode_dist': '', 'profit_center': ''
                        })
                
                log(f"  Found {len(existing_materials)} existing materials")

        # NEW LOGIC: Map unknown storages to GS00
        df_mb51_filtered = df_mb51.copy()
        log(f"MB51 data before storage mapping: {len(df_mb51_filtered)} rows")
        
        # Define known storages
        known_storages = ['GS00', 'BS00', 'AI00', 'TR00', 'EMPTY_STORAGE']
        
        # Find rows with unknown storage
        unknown_storage_mask = ~df_mb51_filtered['storage'].isin(known_storages)
        unknown_storage_count = unknown_storage_mask.sum()
        
        if unknown_storage_count > 0:
            log(f"Found {unknown_storage_count} rows with unknown storage")
            
            # Show unique unknown storages
            unknown_storages = df_mb51_filtered[unknown_storage_mask]['storage'].unique()
            log(f"  Unknown storages: {list(unknown_storages)[:10]}")  # Show first 10
            
            # Map unknown storages to GS00
            df_mb51_filtered.loc[unknown_storage_mask, 'storage'] = 'GS00'
            log(f"  → Mapped all unknown storages to GS00")
        
        log(f"After storage mapping: {len(df_mb51_filtered)} rows (all included)")
        
        # Get unique plants from MB51
        df_mb51_filtered['plant_clean'] = df_mb51_filtered['plant'].astype(str).str.strip().str.upper()
        mb51_plants = set(df_mb51_filtered['plant_clean'].unique())
        log(f"Unique plants in MB51: {len(mb51_plants)}")

        # Group MB51 by (Material, Plant, Storage, Movement Type, Movement Type Text)
        log("=== Grouping MB51 by exact combination ===")
        log("Grouping by: (Material, Plant, Storage, Movement Type, Movement Type Text)")
        
        grouped_mb51 = df_mb51_filtered.groupby(
            ['material', 'plant_clean', 'storage', 'mv_type', 'mv_text'],
            dropna=False
        ).agg({'amount': 'sum'}).reset_index()
        
        log(f"  Grouped MB51: {len(grouped_mb51)} unique combinations")
        
        neg_after_group = (grouped_mb51['amount'] < 0).sum()
        pos_after_group = (grouped_mb51['amount'] > 0).sum()
        sum_after_group = grouped_mb51['amount'].sum()
        
        log(f"  After groupby - Positive: {pos_after_group}, Negative: {neg_after_group}")
        log(f"  After groupby - Total sum: {sum_after_group:,.2f}")
        
        # Map mv_text to mv_grouping
        log("Mapping mv_text to mv_grouping...")
        grouped_mb51['mv_grouping'] = grouped_mb51['mv_text'].map(mv_text_to_grouping)
        
        # Count how many were mapped
        mapped_count = grouped_mb51['mv_grouping'].notna().sum()
        unmapped_count = grouped_mb51['mv_grouping'].isna().sum()
        
        log(f"  Mapped: {mapped_count}/{len(grouped_mb51)}")
        log(f"  Unmapped: {unmapped_count}/{len(grouped_mb51)}")
        
        if unmapped_count > 0:
            log(f"  === UNMAPPED mv_text DETAILS ===")
            unmapped_df = grouped_mb51[grouped_mb51['mv_grouping'].isna()].copy()
            
            # Group by mv_text to see frequency
            unmapped_summary = unmapped_df.groupby(['mv_text', 'mv_type', 'storage']).agg({
                'amount': 'sum',
                'material': 'count'
            }).reset_index()
            unmapped_summary.columns = ['mv_text', 'mv_type', 'storage', 'total_amount', 'count']
            unmapped_summary = unmapped_summary.sort_values('total_amount', ascending=False)
            
            log(f"  Total unmapped combinations: {len(unmapped_summary)}")
            log(f"  All unmapped movements (sorted by total_amount):")
            for idx, row in unmapped_summary.iterrows():
                log(f"    mv_text='{row['mv_text']}', mv_type={row['mv_type']}, storage={row['storage']}, amount={row['total_amount']:,.2f}, occurrences={row['count']}")
        
        # Determine target column
        log("Determining target columns...")
        
        def get_target_column(row):
            storage = row['storage']
            mv_grouping = row['mv_grouping']
            mv_type = row['mv_type']
            
            # Handle empty storage 641/642 separately
            if storage == 'EMPTY_STORAGE':
                if mv_type == '641':
                    return 'BC'
                elif mv_type == '642':
                    return 'BD'
            
            # Use storage + mv_grouping mapping
            if pd.notna(mv_grouping):
                key = (storage, mv_grouping)
                return storage_grouping_to_column.get(key, None)
            
            return None
        
        grouped_mb51['target_column'] = grouped_mb51.apply(get_target_column, axis=1)
        
        # Count how many have target columns
        has_target = grouped_mb51['target_column'].notna().sum()
        no_target = grouped_mb51['target_column'].isna().sum()
        
        log(f"  Has target column: {has_target}/{len(grouped_mb51)}")
        log(f"  No target column: {no_target}/{len(grouped_mb51)}")
        
        if no_target > 0:
            log(f"  === NO TARGET COLUMN DETAILS ===")
            no_target_df = grouped_mb51[grouped_mb51['target_column'].isna()].copy()
            
            # Group by combination to see frequency
            no_target_summary = no_target_df.groupby(['storage', 'mv_grouping', 'mv_type', 'mv_text']).agg({
                'amount': 'sum',
                'material': 'count'
            }).reset_index()
            no_target_summary.columns = ['storage', 'mv_grouping', 'mv_type', 'mv_text', 'total_amount', 'count']
            no_target_summary = no_target_summary.sort_values('total_amount', ascending=False)
            
            log(f"  Total combinations without target: {len(no_target_summary)}")
            log(f"  All combinations without target column (sorted by amount):")
            for idx, row in no_target_summary.iterrows():
                log(f"    Storage={row['storage']}, mv_grouping={row['mv_grouping']}, mv_type={row['mv_type']}, mv_text='{row['mv_text']}', amount={row['total_amount']:,.2f}, count={row['count']}")
        
        # Create lookup: (material, plant, target_column) -> amount
        log("Creating lookup dictionary...")
        mb51_lookup = {}
        
        for _, row in grouped_mb51.iterrows():
            if pd.notna(row['target_column']):
                key = (row['material'], row['plant_clean'], row['target_column'])
                
                # If key exists, add to it (shouldn't happen, but just in case)
                if key in mb51_lookup:
                    mb51_lookup[key] += row['amount']
                else:
                    mb51_lookup[key] = row['amount']
        
        log(f"  Created lookup with {len(mb51_lookup)} keys")
        
        # Verify totals per column
        log(f"  === Totals per target column ===")
        column_totals = defaultdict(float)
        for key, amount in mb51_lookup.items():
            target_col = key[2]
            column_totals[target_col] += amount
        
        for col in sorted(column_totals.keys()):
            log(f"    {col}: {column_totals[col]:,.2f}")
        
        # Verify 641/642 specifically
        bc_total = column_totals.get('BC', 0)
        bd_total = column_totals.get('BD', 0)
        log(f"  BC (641) total: {bc_total:,.2f}")
        log(f"  BD (642) total: {bd_total:,.2f}")

        # Merge materials
        log(f"Merging materials from main file and MB51")
        
        main_file_plants = set()
        if len(existing_materials) > 0:
            main_file_plants = set(m['plant'] for m in existing_materials)
        
        all_materials = existing_materials.copy()
        existing_set = set(f"{m['material']}|{m['plant']}" for m in existing_materials)
        
        mb51_materials = df_mb51_filtered.groupby(
            ['area', 'plant_clean', 'kode_dist', 'profit_center', 'material'], 
            dropna=False
        ).size().reset_index(name='count')
        
        new_materials_added = 0
        
        for _, mb_row in mb51_materials.iterrows():
            plant_normalized = str(mb_row['plant_clean']).strip().upper()
            material = str(mb_row['material']).strip()
            key = f"{material}|{plant_normalized}"
            
            if key in existing_set:
                continue
            
            if len(main_file_plants) > 0 and plant_normalized not in main_file_plants:
                continue
            
            all_materials.append({
                'material': material, 
                'plant': plant_normalized,
                'area': mb_row['area'], 
                'kode_dist': mb_row['kode_dist'],
                'profit_center': mb_row['profit_center']
            })
            new_materials_added += 1
        
        log(f"  Existing materials: {len(existing_materials)}")
        log(f"  New materials: {new_materials_added}")
        log(f"  Total: {len(all_materials)}")

        if len(all_materials) == 0:
            raise ValueError("No materials found")

        grouped_materials = pd.DataFrame(all_materials)

        # Material descriptions
        log("Loading material descriptions...")
        material_desc_map = {}
        
        if 'Output Report INV ARUS BARANG' in sheets_dict:
            try:
                df_out = sheets_dict['Output Report INV ARUS BARANG']
                if df_out.shape[0] > 8 and df_out.shape[1] >= 7:
                    for idx in range(7, len(df_out)):
                        try:
                            row = df_out.iloc[idx]
                            mat_key = str(row.iloc[5]).strip() if len(row) > 5 else ''
                            mat_desc = str(row.iloc[6]).strip() if len(row) > 6 else ''
                            if mat_key and mat_key != 'nan' and mat_desc and mat_desc != 'nan':
                                material_desc_map[mat_key] = mat_desc
                        except:
                            continue
                    log(f"  Loaded {len(material_desc_map)} from main file")
            except Exception as e:
                log(f"  Warning: {str(e)}")
        
        if 'material_desc_mb51' in df_mb51.columns:
            desc_df = df_mb51[['material', 'material_desc_mb51']].dropna()
            desc_df = desc_df[desc_df['material_desc_mb51'].astype(str).str.strip() != '']
            desc_df = desc_df.drop_duplicates('material', keep='first')
            
            added_from_mb51 = 0
            for mat, desc in zip(desc_df['material'], desc_df['material_desc_mb51']):
                if mat not in material_desc_map:
                    material_desc_map[mat] = desc
                    added_from_mb51 += 1
            
            log(f"  Added {added_from_mb51} from MB51")
        
        log(f"  Total: {len(material_desc_map)} descriptions")

        # Create workbook
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Output_Report_INV_ARUS_BARANG_{timestamp}.xlsx"
        output_dir = os.path.join("assets", "exports")
        ensure_dir(output_dir)
        output_path = os.path.join(output_dir, filename)

        log("Creating Excel workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Output Report INV ARUS BARANG"
        center = Alignment(horizontal="center", vertical="center")

        # HEADER
        ws["F1"], ws["F2"], ws["F3"], ws["F4"], ws["F5"], ws["F7"] = "Nama Area", "Plant", "Kode Dist", "Profit Center", "Periode", "Material"
        
        if not grouped_materials.empty:
            first_row = grouped_materials.iloc[0]
            ws["G1"], ws["G2"], ws["G3"], ws["G4"], ws["G5"] = first_row['area'], first_row['plant'], first_row['kode_dist'], first_row['profit_center'], bulan_only
        
        ws["G7"] = "Material Description"
        ws["A8"], ws["B8"], ws["C8"], ws["D8"], ws["E8"], ws["F8"] = "Nama Area", "Plant", "Kode Dist", "Profit Center", "Periode", "source data"
        
        # Row 8 labels - UPDATED for new column layout
        ws["R8"], ws["S8"], ws["T8"], ws["U8"], ws["V8"], ws["W8"], ws["X8"], ws["Y8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["AB8"], ws["AC8"], ws["AD8"], ws["AE8"], ws["AF8"], ws["AG8"] = "DTB", "BPPR", "LBP", "LBP", "ALIH STATUS", "Pemusnahan"
        ws["AJ8"], ws["AK8"], ws["AL8"], ws["AM8"], ws["AN8"], ws["AO8"], ws["AP8"], ws["AQ8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["AT8"], ws["AU8"], ws["AV8"], ws["AW8"], ws["AX8"], ws["AY8"], ws["AZ8"] = "DTB", "BPPR", "LBP", "LBP", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["BC8"], ws["BD8"] = "641", "642"

        ws.merge_cells("H4:M4")
        ws["H4"] = f"SALDO AWAL {bulan} {tahun}"
        ws["H4"].alignment = center
        
        ws.merge_cells("H5:J5")
        ws["H5"] = f"SALDO AWAL {prev_month} {prev_year}"
        ws["H5"].alignment = center
        
        ws.merge_cells("K5:M5")
        ws["K5"] = "SAP - MB5B"
        ws["K5"].alignment = center
        ws["N5"] = "DIFF"
        ws["N5"].alignment = center

        headers_6 = ["GS", "BS", "Grand Total", "GS", "BS", "Grand Total", "GS", "BS", "Grand Total"]
        for i, label in enumerate(headers_6, start=8):
            ws.cell(row=6, column=i, value=label).alignment = center

        for col in range(8, 17):
            ws.cell(row=7, column=col, value="S.Aw").alignment = center

        ws["R1"] = "ctrl balance MB51"
        ws.merge_cells("R5:BE5")
        ws["R5"] = "SAP - MB51"
        ws["R5"].alignment = center

        ws.merge_cells("R6:Z6")
        ws["R6"] = "GS00"
        ws["R6"].alignment = center

        gs00_movements = [
            ("R", "Terima Barang"), ("S", "Retur Beli"), ("T", "Penjualan"),
            ("U", "Retur Jual"), ("V", "Intra Gudang Masuk"), ("W", "Intra Gudang"),
            ("X", "Transfer Stock"), ("Y", "Pemusnahan"), ("Z", "Adjustment")
        ]
        for col, label7 in gs00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        ws.merge_cells("AB6:AH6")
        ws["AB6"] = "BS00"
        ws["AB6"].alignment = center
        bs00_movements = [
            ("AB", "Terima Barang"), ("AC", "Retur Beli"), ("AD", "Penjualan"),
            ("AE", "Retur Jual"), ("AF", "Transfer Stock"), ("AG", "Pemusnahan"), ("AH", "Adjustment"), ("AI", "Intra Gudang Masuk")
        ]
        for col, label7 in bs00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        # UPDATED: AI00 now has 9 columns (added "Intra Gudang Masuk")
        ws.merge_cells("AJ6:AR6")
        ws["AJ6"] = "AI00"
        ws["AJ6"].alignment = center
        ai00_movements = [
            ("AJ", "Terima Barang"), ("AK", "Retur Beli"), ("AL", "Penjualan"),
            ("AM", "Retur Jual"), ("AN", "Intra Gudang Masuk"), ("AO", "Intra Gudang"),
            ("AP", "Transfer Stock"), ("AQ", "Pemusnahan"), ("AR", "Adjustment")
        ]
        for col, label7 in ai00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        # UPDATED: TR00 shifted to AT-BA
        ws.merge_cells("AT6:BA6")
        ws["AT6"] = "TR00"
        ws["AT6"].alignment = center
        tr00_movements = [
            ("AT", "Terima Barang"), ("AU", "Retur Beli"), ("AV", "Penjualan"),
            ("AV", "Retur Jual"), ("AX", "Intra Gudang"), ("AY", "Transfer Stock"),
            ("AZ", "Pemusnahan"), ("BA", "Adjustment")
        ]
        for col, label7 in tr00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        # UPDATED: 641/642 shifted to BC-BE
        ws.merge_cells("BC6:BE6")
        ws["BC6"] = "641 dan 642 tanpa sloc"
        ws["BC6"].alignment = center
        ws["BC7"], ws["BD7"], ws["BE7"] = "Intra Gudang", "Intra Gudang", "CEK"
        ws["BF3"], ws["BF4"] = "-->stock in transit", "jika selisih cek ke MB5T"

        # UPDATED: END STOCK shifted
        ws.merge_cells("BH4:BM4")
        ws["BH4"] = f"END STOCK {prev_month} {prev_year}"
        ws["BH4"].alignment = center
        ws.merge_cells("BH5:BJ5")
        ws["BH5"] = "SALDO AKHIR"
        ws["BH5"].alignment = center
        ws.merge_cells("BK5:BM5")
        ws["BK5"] = "SAP - MB5B"
        ws["BK5"].alignment = center
        ws["BN5"] = "DIFF"
        ws["BN5"].alignment = center

        ws["BH6"], ws["BI6"], ws["BJ6"] = "GS00", "BS00", "Grand Total"
        ws["BK6"], ws["BL6"], ws["BM6"] = "GS", "BS", "Grand Total"
        ws["BN6"], ws["BO6"], ws["BP6"] = "GS", "BS", "Grand Total"

        for col in range(60, 69):
            ws.cell(row=6, column=col).alignment = center
            ws.cell(row=7, column=col, value="S.Ak").alignment = center

        ws["BQ7"] = "CEK SELISIH VS BULAN LALU"
        ws["BR7"] = "kalo ada selisih atas inputan LOG1, LOG2 -> konfirmasi pa Reza utk diselesaikan"

        ws.merge_cells("BS5:BU5")
        ws["BS5"] = "STOCK - EDS"
        ws["BS5"].alignment = center
        ws["BV5"] = "DIFF"
        ws["BV5"].alignment = center

        ws["BS6"], ws["BT6"], ws["BU6"] = "GS", "BS", "Grand Total"
        ws["BV6"], ws["BW6"], ws["BX6"] = "GS", "BS", "Grand Total"

        for col in range(71, 77):
            ws.cell(row=6, column=col).alignment = center
            ws.cell(row=7, column=col, value="S.Ak").alignment = center

        # BODY CALCULATION
        log("Calculating body rows...")
        write_row = 9
        totals = defaultdict(float)
        
        num_materials = len(grouped_materials)
        
        # UPDATED: List of all target columns with new layout
        all_target_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                              "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI",
                              "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
                              "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA",
                              "BC", "BD"]
        
        for idx in range(num_materials):
            if idx % 100 == 0:
                log(f"  Processing {idx}/{num_materials}")
            
            mat_row = grouped_materials.iloc[idx]
            area = str(mat_row['area'])
            plant = str(mat_row['plant']).strip().upper()
            kode_dist = str(mat_row['kode_dist'])
            profit_center = str(mat_row['profit_center'])
            material = str(mat_row['material'])
            material_desc = material_desc_map.get(material, "")
            
            plant_exists_in_mb51 = plant in mb51_plants

            # Write basic info
            ws.cell(row=write_row, column=1, value=area)
            ws.cell(row=write_row, column=2, value=plant)
            ws.cell(row=write_row, column=3, value=kode_dist)
            ws.cell(row=write_row, column=4, value=profit_center)
            ws.cell(row=write_row, column=5, value=bulan_only)
            ws.cell(row=write_row, column=6, value=material)
            ws.cell(row=write_row, column=7, value=material_desc)

            # SALDO AWAL
            h9 = sheet_cache.get_saldo_awal(material, plant, "GS")
            i9 = sheet_cache.get_saldo_awal(material, plant, "BS")
            j9_formula = f"=H{write_row}+I{write_row}"
            
            ws.cell(row=write_row, column=8, value=h9)
            ws.cell(row=write_row, column=9, value=i9)
            ws.cell(row=write_row, column=10, value=j9_formula)

            k9 = sheet_cache.get_mb5b_awal(material, plant, "GS")
            l9 = sheet_cache.get_mb5b_awal(material, plant, "BS")
            m9_formula = f"=SUM(K{write_row}:L{write_row})"
            
            ws.cell(row=write_row, column=11, value=k9)
            ws.cell(row=write_row, column=12, value=l9)
            ws.cell(row=write_row, column=13, value=m9_formula)

            n9_formula = f"=H{write_row}-K{write_row}"
            o9_formula = f"=I{write_row}-L{write_row}"
            p9_formula = f"=N{write_row}+O{write_row}"
            
            ws.cell(row=write_row, column=14, value=n9_formula)
            ws.cell(row=write_row, column=15, value=o9_formula)
            ws.cell(row=write_row, column=16, value=p9_formula)

            # Write all columns from lookup
            if plant_exists_in_mb51:
                for target_col in all_target_columns:
                    lookup_key = (material, plant, target_col)
                    amount = mb51_lookup.get(lookup_key, 0.0)
                    
                    ws.cell(row=write_row, column=get_column_index(target_col), value=amount)
                    totals[target_col] += amount
            else:
                # Plant not in MB51, write zeros
                for target_col in all_target_columns:
                    ws.cell(row=write_row, column=get_column_index(target_col), value=0)

            # UPDATED: BE formula (check) - V vs BC vs BD
            be9_formula = f"=V{write_row}-BC{write_row}-BD{write_row}"
            ws.cell(row=write_row, column=get_column_index("BE"), value=be9_formula)

            # UPDATED: END STOCK formulas (shifted to BH, BI, BJ)
            bh9_formula = f"=H{write_row}+SUM(R{write_row}:Z{write_row})+SUM(AJ{write_row}:BA{write_row})"
            bi9_formula = f"=I{write_row}+SUM(AB{write_row}:AH{write_row})"
            bj9_formula = f"=BH{write_row}+BI{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BH"), value=bh9_formula)
            ws.cell(row=write_row, column=get_column_index("BI"), value=bi9_formula)
            ws.cell(row=write_row, column=get_column_index("BJ"), value=bj9_formula)

            # SAP - MB5B (shifted to BK, BL, BM)
            bk9 = sheet_cache.get_mb5b(material, plant, "GS")
            bl9 = sheet_cache.get_mb5b(material, plant, "BS")
            bm9_formula = f"=SUM(BK{write_row}:BL{write_row})"
            
            ws.cell(row=write_row, column=get_column_index("BK"), value=bk9)
            ws.cell(row=write_row, column=get_column_index("BL"), value=bl9)
            ws.cell(row=write_row, column=get_column_index("BM"), value=bm9_formula)

            # DIFF (shifted to BN, BO, BP)
            bn9_formula = f"=BH{write_row}-BK{write_row}"
            bo9_formula = f"=BI{write_row}-BL{write_row}"
            bp9_formula = f"=BN{write_row}+BO{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BN"), value=bn9_formula)
            ws.cell(row=write_row, column=get_column_index("BO"), value=bo9_formula)
            ws.cell(row=write_row, column=get_column_index("BP"), value=bp9_formula)

            bq9_formula = f"=P{write_row}-BP{write_row}"
            ws.cell(row=write_row, column=get_column_index("BQ"), value=bq9_formula)

            # STOCK - EDS (shifted to BS, BT, BU)
            bs9 = sheet_cache.get_eds(material, plant, "GS")
            bt9 = sheet_cache.get_eds(material, plant, "BS")
            bu9_formula = f"=BS{write_row}+BT{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BS"), value=bs9)
            ws.cell(row=write_row, column=get_column_index("BT"), value=bt9)
            ws.cell(row=write_row, column=get_column_index("BU"), value=bu9_formula)

            bv9_formula = f"=BK{write_row}-BS{write_row}"
            bw9_formula = f"=BL{write_row}-BT{write_row}"
            bx9_formula = f"=BV{write_row}+BW{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BV"), value=bv9_formula)
            ws.cell(row=write_row, column=get_column_index("BW"), value=bw9_formula)
            ws.cell(row=write_row, column=get_column_index("BX"), value=bx9_formula)

            write_row += 1

        log(f"Total rows written: {write_row - 9}")

        # Write formulas
        log("Writing formulas...")
        last_row = write_row - 1
        
        # UPDATED: sum_columns with new layout
        sum_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                      "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"
                      "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR",
                      "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA",
                      "BC", "BD", "BE"]
        
        for col in sum_columns:
            ws[f"{col}3"] = f"=SUM({col}9:{col}{last_row})"
        
        # S1 - ctrl balance
        if len(main_file_plants) > 0:
            mb51_for_s1 = df_mb51_filtered[df_mb51_filtered['plant_clean'].isin(main_file_plants)]
            mb51_total_amount = mb51_for_s1['amount'].sum()
        else:
            mb51_total_amount = df_mb51_filtered['amount'].sum()
        
        sum_r3_bc3 = sum([totals.get(col, 0) for col in sum_columns])
        s1_value = round(mb51_total_amount - sum_r3_bc3, 2)
        
        ws["S1"] = s1_value
        log(f"  S1 = {s1_value:.2f}")
        log(f"  S1 breakdown: MB51_total={mb51_total_amount:,.2f} - Columns_sum={sum_r3_bc3:,.2f}")
        
        # UPDATED: AY2 formula (Transfer Stock total) - X, AF, AP, AY
        ws["AY2"] = "=X3+AF3+AP3+AY3"
        
        # BM2 calculation
        sum_bm = 0.0
        for row in range(9, write_row):
            cell_val = ws.cell(row=row, column=get_column_index("BM")).value
            if isinstance(cell_val, (int, float)):
                sum_bm += cell_val
        
        sum_mb5b_pq = 0.0
        if '13. MB5B' in sheets_dict:
            df_mb5b_sheet = sheets_dict['13. MB5B']
            try:
                if df_mb5b_sheet.shape[1] > 16:
                    sum_p = pd.to_numeric(df_mb5b_sheet.iloc[:, 15], errors='coerce').fillna(0).sum()
                    sum_q = pd.to_numeric(df_mb5b_sheet.iloc[:, 16], errors='coerce').fillna(0).sum()
                    sum_mb5b_pq = sum_p + sum_q
            except Exception as e:
                log(f"  Warning: {str(e)}")
        
        bm2_value = round(sum_bm - sum_mb5b_pq, 2)
        ws["BM2"] = bm2_value
        log(f"  BM2 = {bm2_value:.2f}")

        # Formatting
        log("Formatting...")
        for i in range(1, 80):
            ws.column_dimensions[get_column_letter(i)].width = 12
        
        ws.column_dimensions['Q'].width = 2
        ws.column_dimensions['AA'].width = 2
        ws.column_dimensions['AI'].width = 2
        ws.column_dimensions['AS'].width = 2
        ws.column_dimensions['BB'].width = 2
        ws.column_dimensions['BG'].width = 4

        ws.freeze_panes = "H9"

        for row in range(9, write_row):
            for col in range(8, 80):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        
        for row in [2, 3]:
            for col in range(18, 80):
                ws.cell(row=row, column=col).number_format = '#,##0'

        # Save
        log(f"Saving workbook...")
        wb.save(output_path)
        
        if not os.path.exists(output_path):
            raise Exception(f"File was not created")
        
        file_size = os.path.getsize(output_path)
        log(f"✓ File created: {file_size:,} bytes")
        
        result = {
            "success": True,
            "output_path": output_path,
            "rows_written": write_row - 9,
            "report_month": f"{bulan} {tahun}",
            "total_materials": len(grouped_materials),
            "file_size": file_size,
            "timestamp": timestamp,
            "unmapped_count": int(unmapped_count),
            "no_target_count": int(no_target)
        }
        
        print(json.dumps(result))
        sys.stdout.flush()
        log("✓ Report completed successfully with NEW AI00 column!")

    except Exception as e:
        tb = traceback.format_exc()
        log(f"ERROR: {str(e)}")
        log(f"Traceback:\n{tb}")
        
        error_result = {
            "success": False,
            "error": str(e),
            "trace": tb
        }
        
        print(json.dumps(error_result))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()