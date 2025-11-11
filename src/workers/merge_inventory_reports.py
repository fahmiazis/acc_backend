# merge_inventory_reports.py
# Worker Python untuk menggabungkan multiple "Output Report INV ARUS BARANG" menjadi satu file
# Input: List of Excel file paths (hasil generate sebelumnya)
# Output: 1 consolidated Excel file dengan semua data + SUM formulas

import sys
import json
import os
import datetime
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def log(msg):
    """Log to stderr so it doesn't interfere with JSON output on stdout"""
    print(f"[merge-worker] {msg}", file=sys.stderr, flush=True)

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def copy_cell_style(source_cell, target_cell):
    """Copy all styling from source cell to target cell"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()

def main():
    try:
        payload = json.load(sys.stdin)
        file_paths = payload.get("file_paths", [])
        
        if not file_paths or len(file_paths) == 0:
            raise ValueError("No file paths provided")
        
        log(f"Received {len(file_paths)} files to merge:")
        for i, fp in enumerate(file_paths, 1):
            log(f"  {i}. {fp}")
        
        # Validate all files exist
        for fp in file_paths:
            if not os.path.exists(fp):
                raise FileNotFoundError(f"File not found: {fp}")
        
        # Create new workbook
        log("Creating consolidated workbook...")
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Output Report INV ARUS BARANG"
        
        # Variables to track data
        current_write_row = 9  # Start writing data from row 9
        header_copied = False
        plant_codes = set()  # Collect all plant codes
        
        # Process each file
        for file_idx, file_path in enumerate(file_paths):
            log(f"\nProcessing file {file_idx + 1}/{len(file_paths)}: {os.path.basename(file_path)}")
            
            # Load workbook
            wb_source = load_workbook(file_path, data_only=False)  # Keep formulas
            
            # Find the sheet (should be "Output Report INV ARUS BARANG")
            sheet_name = "Output Report INV ARUS BARANG"
            if sheet_name not in wb_source.sheetnames:
                log(f"  WARNING: Sheet '{sheet_name}' not found, trying first sheet")
                ws_source = wb_source.worksheets[0]
            else:
                ws_source = wb_source[sheet_name]
            
            log(f"  Source sheet: {ws_source.title}, dimensions: {ws_source.dimensions}")
            
            # Collect plant codes for G2 (do this BEFORE header copy check)
            source_plant = ws_source["G2"].value
            if source_plant and str(source_plant).strip() and str(source_plant).strip() != 'nan':
                plant_codes.add(str(source_plant).strip())
                log(f"  Found plant code: {source_plant}")
            
            # Copy header (rows 1-8) only from first file
            if not header_copied:
                log("  Copying header (rows 1-8)...")
                
                for row_idx in range(1, 9):  # Rows 1-8
                    for col_idx in range(1, ws_source.max_column + 1):
                        source_cell = ws_source.cell(row=row_idx, column=col_idx)
                        target_cell = ws_output.cell(row=row_idx, column=col_idx)
                        
                        # Copy value
                        target_cell.value = source_cell.value
                        
                        # Copy style
                        copy_cell_style(source_cell, target_cell)
                
                # Copy column widths
                for col_letter in ws_source.column_dimensions:
                    if col_letter in ws_source.column_dimensions:
                        ws_output.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width
                
                # Copy merged cells in header
                for merged_range in ws_source.merged_cells.ranges:
                    if merged_range.min_row <= 8:  # Only header merges
                        ws_output.merge_cells(str(merged_range))
                
                # CUSTOM HEADER for merged report
                ws_output["G1"].value = "Merge Report"
                ws_output["G2"].value = ""  # Will be filled with plant codes later
                ws_output["G3"].value = "-"
                ws_output["G4"].value = "-"
                
                header_copied = True
                log("  Header copied successfully")
                log("  Custom header set: G1='Merge Report', G3='-', G4='-'")
            
            # Copy data rows (from row 9 onwards)
            data_row_count = 0
            source_start_row = 9
            
            # Find last row with data in source
            source_last_row = ws_source.max_row
            
            # Count actual data rows (skip empty rows)
            for row_idx in range(source_start_row, source_last_row + 1):
                # Check if row has data (check column F - Material)
                material_cell = ws_source.cell(row=row_idx, column=6)
                if material_cell.value and str(material_cell.value).strip() and str(material_cell.value).strip() != 'nan':
                    data_row_count += 1
            
            log(f"  Found {data_row_count} data rows (rows {source_start_row} to {source_last_row})")
            
            if data_row_count == 0:
                log("  WARNING: No data rows found, skipping this file")
                wb_source.close()
                continue
            
            # Copy data rows
            log(f"  Copying data to output starting at row {current_write_row}...")
            rows_copied = 0
            
            for row_idx in range(source_start_row, source_last_row + 1):
                # Check if row has data
                material_cell = ws_source.cell(row=row_idx, column=6)
                if not material_cell.value or str(material_cell.value).strip() == '' or str(material_cell.value).strip() == 'nan':
                    continue
                
                # Copy all columns
                for col_idx in range(1, ws_source.max_column + 1):
                    source_cell = ws_source.cell(row=row_idx, column=col_idx)
                    target_cell = ws_output.cell(row=current_write_row, column=col_idx)
                    
                    # Copy value/formula
                    if source_cell.data_type == 'f':  # Formula
                        # For formulas, just copy the value (data_only mode would be better but we need formulas in header)
                        target_cell.value = source_cell.value
                    else:
                        target_cell.value = source_cell.value
                    
                    # Copy style
                    copy_cell_style(source_cell, target_cell)
                
                current_write_row += 1
                rows_copied += 1
            
            log(f"  Copied {rows_copied} data rows")
            
            # Close source workbook
            wb_source.close()
        
        total_data_rows = current_write_row - 9
        log(f"\nTotal data rows in consolidated file: {total_data_rows}")
        
        if total_data_rows == 0:
            raise ValueError("No data rows found in any of the source files")
        
        # Update G2 with all plant codes
        if plant_codes:
            plant_codes_str = ", ".join(sorted(plant_codes))
            ws_output["G2"].value = plant_codes_str
            log(f"Updated G2 with plant codes: {plant_codes_str}")
        
        # Apply freeze panes: Freeze rows 1-8 and columns A-G
        # Freeze point is at H9 (first unfrozen cell)
        ws_output.freeze_panes = "H9"
        log("Applied freeze panes at H9 (rows 1-8 and columns A-G frozen)")
        
        # Update SUM formulas in row 3 (R3:BD3)
        log("Updating SUM formulas in row 3...")
        last_data_row = current_write_row - 1
        
        sum_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                      "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                      "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ",
                      "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                      "BB", "BC", "BD"]
        
        for col_letter in sum_columns:
            formula = f"=SUM({col_letter}9:{col_letter}{last_data_row})"
            ws_output[f"{col_letter}3"].value = formula
            ws_output[f"{col_letter}3"].number_format = '#,##0'
        
        log(f"  Updated {len(sum_columns)} SUM formulas")
        
        # Update other formulas that depend on row numbers
        log("Updating other formulas...")
        
        # S1 formula
        ws_output["S1"].value = "='5. MB51'!M1-SUM('Output Report INV ARUS BARANG'!R3:BB3)"
        
        # AX2 formula
        ws_output["AX2"].value = "=X3+AF3+AO3+AX3"
        
        # BL2 formula
        ws_output["BL2"].value = f"=SUM(BL9:BL{last_data_row})-SUM('13. MB5B'!P:Q)"
        
        log("  Formulas updated")
        
        # Save output file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Consolidated_Report_INV_ARUS_BARANG_{timestamp}.xlsx"
        output_dir = os.path.join("assets", "exports")
        ensure_dir(output_dir)
        output_path = os.path.join(output_dir, filename)
        
        log(f"Saving consolidated file to: {output_path}")
        wb_output.save(output_path)
        wb_output.close()
        
        if not os.path.exists(output_path):
            raise Exception(f"File was not created at {output_path}")
        
        file_size = os.path.getsize(output_path)
        log(f"File created successfully, size: {file_size} bytes")
        
        result = {
            "success": True,
            "output_path": output_path,
            "total_files_merged": len(file_paths),
            "total_data_rows": total_data_rows,
            "file_size": file_size,
            "timestamp": timestamp
        }
        
        print(json.dumps(result))
        sys.stdout.flush()
        log("Merge completed successfully!")
        
    except Exception as e:
        tb = traceback.format_exc()
        log(f"Error occurred: {str(e)}")
        log(f"Traceback: {tb}")
        
        error_result = {
            "success": False,
            "error": str(e),
            "trace": tb
        }
        
        print(json.dumps(error_result))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()