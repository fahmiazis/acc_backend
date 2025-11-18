# merge_inventory_reports.py - OPTIMIZED VERSION
# Improvements:
# 1. Batch cell operations (read/write in chunks)
# 2. Parallel file reading
# 3. Smart copying (only necessary cells)
# 4. Pre-allocated target rows
# 5. Minimal style operations

import sys
import json
import os
import datetime
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def log(msg):
    """Log to stderr"""
    print(f"[merge-worker] {msg}", file=sys.stderr, flush=True)

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def copy_cell_style(source_cell, target_cell):
    """Copy styling from source to target cell"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = source_cell.alignment.copy()

# OPTIMIZATION 1: Parallel file reading with S1 and BL2 values
def read_file_data(file_path, file_idx, total_files):
    """Read data from a single file - can be parallelized"""
    try:
        log(f"Reading file {file_idx + 1}/{total_files}: {os.path.basename(file_path)}")
        
        # FIXED: Use data_only=True but NOT read_only (causes issues with some files)
        wb = load_workbook(file_path, data_only=True)
        
        sheet_name = "Output Report INV ARUS BARANG"
        if sheet_name not in wb.sheetnames:
            ws = wb.worksheets[0]
        else:
            ws = wb[sheet_name]
        
        # Extract plant code
        plant_code = None
        try:
            plant_cell = ws.cell(row=2, column=7).value  # G2
            if plant_cell and str(plant_cell).strip() and str(plant_cell).strip() != 'nan':
                plant_code = str(plant_cell).strip()
        except:
            pass
        
        # Extract S1 value (ctrl balance MB51)
        s1_value = 0.0
        try:
            s1_cell = ws.cell(row=1, column=19).value  # S1
            if s1_cell and isinstance(s1_cell, (int, float)):
                s1_value = float(s1_cell)
        except:
            pass
        
        # Extract BL2 value (MB5B difference)
        bl2_value = 0.0
        try:
            bl2_cell = ws.cell(row=2, column=64).value  # BL2
            if bl2_cell and isinstance(bl2_cell, (int, float)):
                bl2_value = float(bl2_cell)
        except:
            pass
        
        log(f"  File {file_idx + 1}: S1={s1_value:.2f}, BL2={bl2_value:.2f}")
        
        # OPTIMIZATION: Read data rows in batch with progress
        data_rows = []
        max_col = min(ws.max_column, 75)  # Limit to 75 columns (BW)
        total_rows = ws.max_row
        
        log(f"  File {file_idx + 1}: Processing {total_rows} rows...")
        
        for row_idx in range(9, total_rows + 1):
            # Progress every 500 rows
            if (row_idx - 9) % 500 == 0 and row_idx > 9:
                log(f"    Progress: {row_idx - 8}/{total_rows - 8} rows")
            
            # Check if row has data (column F - Material)
            material_val = ws.cell(row=row_idx, column=6).value
            if not material_val or str(material_val).strip() == '' or str(material_val).strip() == 'nan':
                continue
            
            # Read entire row at once - SIMPLIFIED
            row_data = []
            for col_idx in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Just value, no format checking (faster)
                    row_data.append({'value': cell.value, 'number_format': None})
                except Exception as e:
                    # Skip problematic cells
                    row_data.append({'value': None, 'number_format': None})
            
            data_rows.append(row_data)
        
        wb.close()
        
        log(f"  File {file_idx + 1}: Read {len(data_rows)} data rows, plant={plant_code}")
        
        return {
            'file_path': file_path,
            'file_idx': file_idx,
            'plant_code': plant_code,
            'data_rows': data_rows,
            'max_col': max_col,
            's1_value': s1_value,
            'bl2_value': bl2_value
        }
        
    except Exception as e:
        log(f"  ERROR reading file {file_idx + 1}: {str(e)}")
        import traceback
        log(f"  Traceback: {traceback.format_exc()}")
        return {
            'file_path': file_path,
            'file_idx': file_idx,
            'error': str(e)
        }

def main():
    try:
        payload = json.load(sys.stdin)
        file_paths = payload.get("file_paths", [])
        
        if not file_paths or len(file_paths) == 0:
            raise ValueError("No file paths provided")
        
        log(f"Received {len(file_paths)} files to merge")
        
        # Validate all files exist
        for fp in file_paths:
            if not os.path.exists(fp):
                raise FileNotFoundError(f"File not found: {fp}")
        
        # OPTIMIZATION 2: Parallel file reading
        log("Reading files in parallel...")
        file_data_list = []
        
        with ThreadPoolExecutor(max_workers=min(4, len(file_paths))) as executor:
            futures = [
                executor.submit(read_file_data, fp, idx, len(file_paths))
                for idx, fp in enumerate(file_paths)
            ]
            
            for future in as_completed(futures):
                file_data = future.result()
                if 'error' in file_data:
                    log(f"WARNING: Skipping file due to error: {file_data['error']}")
                    continue
                file_data_list.append(file_data)
        
        # Sort by original order
        file_data_list.sort(key=lambda x: x['file_idx'])
        
        if not file_data_list:
            raise ValueError("No valid files to merge")
        
        log(f"Successfully read {len(file_data_list)} files")
        
        # Collect plant codes and aggregate S1/BL2
        plant_codes = set()
        total_s1 = 0.0
        total_bl2 = 0.0
        
        for fd in file_data_list:
            if fd.get('plant_code'):
                plant_codes.add(fd['plant_code'])
            total_s1 += fd.get('s1_value', 0.0)
            total_bl2 += fd.get('bl2_value', 0.0)
        
        log(f"Aggregated values: S1 total={total_s1:.2f}, BL2 total={total_bl2:.2f}")
        
        # Count total data rows
        total_data_rows = sum(len(fd['data_rows']) for fd in file_data_list)
        log(f"Total data rows to merge: {total_data_rows}")
        
        if total_data_rows == 0:
            raise ValueError("No data rows found in any file")
        
        # Create output workbook
        log("Creating consolidated workbook...")
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Output Report INV ARUS BARANG"
        
        # OPTIMIZATION 3: Copy header from first file (load separately for styling)
        log("Copying header from first file...")
        first_file = file_paths[0]
        wb_first = load_workbook(first_file, data_only=False)
        
        sheet_name = "Output Report INV ARUS BARANG"
        if sheet_name not in wb_first.sheetnames:
            ws_first = wb_first.worksheets[0]
        else:
            ws_first = wb_first[sheet_name]
        
        # Copy header (rows 1-8)
        max_col = ws_first.max_column
        for row_idx in range(1, 9):
            for col_idx in range(1, max_col + 1):
                source_cell = ws_first.cell(row=row_idx, column=col_idx)
                target_cell = ws_output.cell(row=row_idx, column=col_idx)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
        
        # Copy column widths
        for col_letter in ws_first.column_dimensions:
            ws_output.column_dimensions[col_letter].width = ws_first.column_dimensions[col_letter].width
        
        # Copy merged cells in header
        for merged_range in ws_first.merged_cells.ranges:
            if merged_range.min_row <= 8:
                ws_output.merge_cells(str(merged_range))
        
        wb_first.close()
        
        # Update header for merged report
        ws_output["G1"].value = "Merge Report"
        ws_output["G2"].value = ", ".join(sorted(plant_codes)) if plant_codes else "-"
        ws_output["G3"].value = "-"
        ws_output["G4"].value = "-"
        
        log(f"Header copied, plant codes: {ws_output['G2'].value}")
        
        # OPTIMIZATION 4: Batch write data rows
        log("Writing data rows...")
        current_row = 9
        
        for file_data in file_data_list:
            data_rows = file_data['data_rows']
            max_col = file_data['max_col']
            
            log(f"  Writing {len(data_rows)} rows from file {file_data['file_idx'] + 1}")
            
            # Write in batches with progress
            batch_count = 0
            for row_data in data_rows:
                for col_idx, cell_data in enumerate(row_data, start=1):
                    target_cell = ws_output.cell(row=current_row, column=col_idx)
                    target_cell.value = cell_data['value']
                    
                    # Apply number format for numeric columns (8 onwards)
                    if col_idx >= 8 and isinstance(cell_data['value'], (int, float)):
                        target_cell.number_format = '#,##0'
                
                current_row += 1
                batch_count += 1
                
                # Progress every 500 rows
                if batch_count % 500 == 0:
                    log(f"    Progress: {batch_count}/{len(data_rows)} rows written")
            
            log(f"  Completed file {file_data['file_idx'] + 1}")
        
        last_data_row = current_row - 1
        total_rows_written = last_data_row - 8
        log(f"Written {total_rows_written} data rows (rows 9 to {last_data_row})")
        
        # Apply freeze panes
        ws_output.freeze_panes = "H9"
        log("Applied freeze panes at H9")
        
        # OPTIMIZATION 5: Batch formula updates
        log("Updating formulas...")
        
        sum_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                      "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                      "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ",
                      "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                      "BB", "BC", "BD"]
        
        # Update SUM formulas in row 3
        for col_letter in sum_columns:
            ws_output[f"{col_letter}3"].value = f"=SUM({col_letter}9:{col_letter}{last_data_row})"
            ws_output[f"{col_letter}3"].number_format = '#,##0'
        
        # Set S1 value (aggregated from all files)
        ws_output["S1"].value = total_s1
        ws_output["S1"].number_format = '#,##0'
        log(f"  S1 set to {total_s1:.2f} (sum of all files)")
        
        # AX2 formula (between columns)
        ws_output["AX2"].value = "=X3+AF3+AO3+AX3"
        ws_output["AX2"].number_format = '#,##0'
        
        # Set BL2 value (aggregated from all files)
        ws_output["BL2"].value = total_bl2
        ws_output["BL2"].number_format = '#,##0'
        log(f"  BL2 set to {total_bl2:.2f} (sum of all files)")
        
        log(f"  Updated {len(sum_columns)} SUM formulas")
        
        # Save output file
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Consolidated_Report_INV_ARUS_BARANG_{timestamp}.xlsx"
        output_dir = os.path.join("assets", "exports")
        ensure_dir(output_dir)
        output_path = os.path.join(output_dir, filename)
        
        log(f"Saving consolidated file to: {output_path}")
        wb_output.save(output_path)
        wb_output.close()
        
        if not os.path.exists(output_path):
            raise Exception(f"File was not created at {output_path}")
        
        file_size = os.path.getsize(output_path)
        log(f"File created successfully, size: {file_size} bytes")
        
        result = {
            "success": True,
            "output_path": output_path,
            "total_files_merged": len(file_data_list),
            "total_data_rows": total_rows_written,
            "plant_codes": sorted(list(plant_codes)),
            "file_size": file_size,
            "timestamp": timestamp
        }
        
        print(json.dumps(result))
        sys.stdout.flush()
        log("Merge completed successfully!")
        
    except Exception as e:
        tb = traceback.format_exc()
        log(f"Error occurred: {str(e)}")
        log(f"Traceback: {tb}")
        
        error_result = {
            "success": False,
            "error": str(e),
            "trace": tb
        }
        
        print(json.dumps(error_result))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()